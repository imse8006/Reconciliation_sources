"""Market configuration: maps market names to ERP system and mapping file.

markets.json schema:
  {
    "MarketName": {
      "erp": "ERPName",           # required: e.g. "Jeeves", "Prophet", "SAP", "AX", "ASW", "Chorus"
      "mapping": "Mapping/..."    # optional: path to attribute mapping Excel file
    }
  }
"""
import json
from pathlib import Path

MARKETS_FILE = Path("markets.json")


def _load() -> dict:
    with open(MARKETS_FILE, encoding="utf-8") as f:
        return json.load(f)


def list_markets() -> list[str]:
    """Return all configured market names."""
    return list(_load().keys())


def get_erp_name(market: str) -> str:
    """Return ERP name for a market (e.g. 'Jeeves', 'Prophet', 'SAP')."""
    cfg = _load()
    if market not in cfg:
        raise ValueError(f"Market '{market}' not in markets.json. Known: {list(cfg)}")
    return cfg[market]["erp"]


def get_mapping_path(market: str) -> Path | None:
    """Return mapping Excel path for a market, or None if not configured."""
    cfg = _load()
    if market not in cfg:
        raise ValueError(f"Market '{market}' not in markets.json.")
    mp = cfg[market].get("mapping")
    return Path(mp) if mp else None


def load_mapping_rows(
    market: str, sheet: str
) -> list[tuple[str | None, str | None, str | None]]:
    """Return list of (stibo_col, erp_col, ct_col) from a mapping sheet.

    erp_col is None when the cell is empty → attribute not available in ERP,
    so attributes reconciliation will compare STIBO vs CT only for that row.
    """
    mp = get_mapping_path(market)
    if mp is None:
        raise FileNotFoundError(f"No mapping file configured for market '{market}'.")
    if not mp.exists():
        raise FileNotFoundError(f"Mapping file not found: {mp}")

    import openpyxl

    wb = openpyxl.load_workbook(mp, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not in {mp.name}. Available: {wb.sheetnames}")
    ws = wb[sheet]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        stibo = row[0] if len(row) > 0 else None
        erp = row[1] if len(row) > 1 else None
        ct = row[2] if len(row) > 2 else None
        # Normalize empty strings to None
        stibo = (str(stibo).strip() or None) if stibo is not None else None
        erp = (str(erp).strip() or None) if erp is not None else None
        ct = (str(ct).strip() or None) if ct is not None else None
        if stibo is not None:
            rows.append((stibo, erp, ct))
    return rows
