"""Product (Range) reconciliation: STIBO (MDM) vs CT vs local ERP.
ERP files are read from ERP/{market}/{date}/. ERP name is read from markets.json."""
import polars as pl
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import hashlib
import json

def load_jeves_data(file_path: str) -> pl.DataFrame:
    """Load JEEVES Product data from sheet 2-EXCELMASTER.

    Product codes are read from column A starting at A3 (row 3).
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["2-EXCELMASTER"]

    data: list[tuple[object]] = []
    for (val,) in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        data.append((val,))

    if not data:
        return pl.DataFrame(schema=["SUPC"])

    return pl.DataFrame(data, schema=["SUPC"], orient="row", infer_schema_length=None)

def load_ct_data(file_path: str) -> pl.DataFrame:
    """Load CT Ekofisk data
    Headers row 6, data starts at B7 (first SUPC)
    """
    if file_path.endswith('.xlsb'):
        try:
            from pyxlsb import open_workbook
        except ModuleNotFoundError:
            raise ModuleNotFoundError(
                "Reading .xlsb files requires pyxlsb. Install with: pip install pyxlsb"
            ) from None

    # Detect file format
    if file_path.endswith('.xlsb'):
        # Read .xlsb file
        with open_workbook(file_path) as wb:
            # Find sheet "Item" or use first sheet
            sheet_name = None
            for name in wb.sheets:
                if name.lower() == 'item':
                    sheet_name = name
                    break
            if sheet_name is None:
                sheet_name = wb.sheets[0]
            
            with wb.get_sheet(sheet_name) as sheet:
                rows = list(sheet.rows())
                
                # Read headers from row 6 (index 5), starting at column B (index 1)
                headers = []
                header_counts = {}  # To handle duplicates
                if len(rows) > 5:
                    header_row = rows[5]  # Row 6 (0-indexed = 5)
                    for idx, cell in enumerate(header_row):
                        if idx >= 1:  # Column B and following (SUPC is in column 2)
                            val = cell.v if cell.v is not None else None
                            if val is None:
                                header_name = f"Col_{idx+1}"
                            else:
                                # Convert to string and clean
                                header_name = str(val).strip() if isinstance(val, str) else str(val)
                            
                            # Handle duplicates
                            if header_name in header_counts:
                                header_counts[header_name] += 1
                                header_name = f"{header_name}_{header_counts[header_name]}"
                            else:
                                header_counts[header_name] = 0
                            
                            headers.append(header_name)
                
                # Read data from row 7 (index 6), column B (index 1)
                data = []
                for row_idx in range(6, len(rows)):
                    row = rows[row_idx]
                    # Check if column B (index 1, SUPC) has a numeric value
                    if len(row) > 1:
                        supc_val = row[1].v
                        # Accept only rows with valid SUPC (numeric)
                        if supc_val is not None and (isinstance(supc_val, (int, float)) or str(supc_val).strip()):
                            row_data = []
                            # Start at column B (index 1)
                            for col_idx in range(1, min(len(row), len(headers) + 1)):
                                cell_val = row[col_idx].v if col_idx < len(row) else None
                                row_data.append(cell_val)
                            # Complete with None if necessary
                            while len(row_data) < len(headers):
                                row_data.append(None)
                            data.append(row_data[:len(headers)])
                
                # Create DataFrame without strict schema to let Polars infer types
                if data:
                    return pl.DataFrame(data, schema=headers, orient="row", infer_schema_length=None)
                else:
                    return pl.DataFrame(schema=headers)
    else:
        # Use openpyxl for .xlsx files
        wb = load_workbook(file_path, data_only=True)
        # Find Product sheet or use active sheet
        ws = None
        for sheet_name in wb.sheetnames:
            if 'product' in sheet_name.lower():
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
        
        # Read headers from row 6, starting at column B
        headers = []
        for cell in ws[6]:
            if cell.column >= 2:  # Column B and following
                if cell.value is not None:
                    headers.append(cell.value)
                else:
                    headers.append(f"Col_{cell.column}")
        
        # Read data from row 7, column B
        data = []
        for row in ws.iter_rows(min_row=7, min_col=2, values_only=True):
            # Check if first column (B) has a value (SUPC)
            if row and row[0] is not None:
                data.append(row[:len(headers)])
        
        return pl.DataFrame(data, schema=headers, orient="row")

def load_stibo_data(file_path: str) -> pl.DataFrame:
    """Load STIBO Product data

    Headers are on row 1. Product codes are read from column C (SUPC) starting at C2.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    supc_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        if str(cell.value).strip().upper() == "SUPC":
            supc_col_idx = idx
            break
    if supc_col_idx is None:
        supc_col_idx = 3  # Column C

    data: list[tuple[object]] = []
    for (val,) in ws.iter_rows(
        min_row=2,
        min_col=supc_col_idx,
        max_col=supc_col_idx,
        values_only=True,
    ):
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        data.append((val,))

    if not data:
        return pl.DataFrame(schema=["SUPC"])

    return pl.DataFrame(data, schema=["SUPC"], orient="row", infer_schema_length=None)


def clean_product_code(value):
    """Clean a product code to remove .0 and normalize format"""
    if value is None:
        return None
    # Convert to string first
    str_val = str(value)
    # If it's a number with .0 at the end, convert to integer then string
    try:
        float_val = float(str_val)
        if float_val.is_integer():
            return str(int(float_val))
        return str_val
    except (ValueError, TypeError):
        return str_val.strip()

def load_prophet_product_data(file_path: str) -> pl.DataFrame:
    """Load Prophet Product data. Headers on row 2, product code column = 'FD Product Code'."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]

    # Find 'FD Product Code' column in row 2
    headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
    col_idx = headers.get("FD Product Code")
    if col_idx is None:
        raise ValueError(
            f"Column 'FD Product Code' not found in row 2 of {file_path}. "
            f"Available: {[h for h in headers if h]}"
        )

    data = []
    for row in range(3, ws.max_row + 1):
        v = ws.cell(row, col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        data.append((v,))
    wb.close()

    if not data:
        return pl.DataFrame(schema=["SUPC"])
    return pl.DataFrame(data, schema=["SUPC"], orient="row", infer_schema_length=None)


def _load_erp_product_data(file_path: str, erp_name: str) -> pl.DataFrame:
    """Dispatch to the right ERP product loader based on erp_name."""
    name = erp_name.lower()
    if name == "jeeves":
        return load_jeves_data(file_path)
    if name == "prophet":
        return load_prophet_product_data(file_path)
    raise NotImplementedError(
        f"No product loader implemented for ERP '{erp_name}'. "
        f"Add one in reconcile_products.py."
    )


def create_range_reconciliation(
    erp_df: pl.DataFrame,
    ct_df: pl.DataFrame,
    stibo_df: pl.DataFrame,
    erp_name: str = "ERP",
) -> pl.DataFrame:
    """Range Reconciliation: list all products with CT / ERP / STIBO columns and X marks."""
    ct_product_col = "SUPC" if "SUPC" in ct_df.columns else ct_df.columns[0]
    stibo_product_col = "SUPC" if "SUPC" in stibo_df.columns else stibo_df.columns[0]
    erp_product_col = "SUPC" if "SUPC" in erp_df.columns else erp_df.columns[0]

    def clean_and_convert(df, col_name):
        return df.select([pl.col(col_name)]).unique().with_columns([
            pl.col(col_name).map_elements(
                lambda x: clean_product_code(x),
                return_dtype=pl.Utf8,
            ).alias("ProductCode")
        ]).select("ProductCode").unique()

    erp_clean = clean_and_convert(erp_df, erp_product_col)
    ct_clean = clean_and_convert(ct_df, ct_product_col)
    stibo_clean = clean_and_convert(stibo_df, stibo_product_col)

    all_products = pl.concat([erp_clean, ct_clean, stibo_clean]).unique("ProductCode")

    erp_list = erp_clean.to_series().to_list()
    ct_list = ct_clean.to_series().to_list()
    stibo_list = stibo_clean.to_series().to_list()

    erp_present_col = f"{erp_name}_present"

    reconciliation = all_products.with_columns([
        pl.col("ProductCode").is_in(ct_list).alias("CT_present"),
        pl.col("ProductCode").is_in(erp_list).alias(erp_present_col),
        pl.col("ProductCode").is_in(stibo_list).alias("STIBO_present"),
    ]).with_columns([
        pl.when(pl.col("CT_present")).then(pl.lit("X")).otherwise(pl.lit("")).alias("CT"),
        pl.when(pl.col(erp_present_col)).then(pl.lit("X")).otherwise(pl.lit("")).alias(erp_name),
        pl.when(pl.col("STIBO_present")).then(pl.lit("X")).otherwise(pl.lit("")).alias("STIBO"),
    ]).with_columns([
        pl.concat_str([
            pl.when(pl.col("CT_present") == False).then(pl.lit("CT")).otherwise(pl.lit("")),
            pl.when(pl.col(erp_present_col) == False).then(pl.lit(erp_name)).otherwise(pl.lit("")),
            pl.when(pl.col("STIBO_present") == False).then(pl.lit("STIBO")).otherwise(pl.lit("")),
        ], separator=", ").str.strip_chars_start(", ").str.strip_chars_end(", ").alias("Absent_from")
    ]).with_columns([
        pl.when(pl.col("Absent_from") == "").then(pl.lit("-")).otherwise(pl.col("Absent_from")).alias("Absent_from")
    ]).select(["ProductCode", "CT", erp_name, "STIBO", "Absent_from"]).sort("ProductCode")

    return reconciliation

def _find_first_file(directory: Path, needle: str) -> Path | None:
    """First file in directory whose name contains needle (case-insensitive)."""
    if not directory.exists() or not directory.is_dir():
        return None
    needle_l = needle.lower()
    for f in sorted(directory.iterdir()):
        if f.is_file() and needle_l in f.name.lower():
            return f
    return None


def get_file_hash(file_path: str) -> str:
    """Calculate MD5 hash of a file to detect changes"""
    hash_md5 = hashlib.md5()
    try:
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()
    except FileNotFoundError:
        return None

def _resolve_product_paths(
    date_folder: str, market: str = "Ekofisk", erp_name: str = "Jeeves"
) -> tuple[Path | None, Path | None, Path | None]:
    """Resolve ERP, CT, STIBO Product file paths. Returns (erp_path, ct_path, stibo_path).
    ERP files are looked up in ERP/{erp_name}/{date}/ (fallback: ERP/{erp_name}/).
    """
    date_folder = date_folder.strip()
    erp_base = Path("ERP") / erp_name
    erp_dir = erp_base / date_folder if (erp_base / date_folder).is_dir() else erp_base
    ct_dir = Path("CT") / date_folder if (Path("CT") / date_folder).is_dir() else Path("CT")
    stibo_dir = Path("STIBO") / date_folder

    erp_path = _find_first_file(erp_dir, "Product")
    ct_path = _find_first_file(ct_dir, "Product") or _find_first_file(Path("CT"), "Product")
    stibo_path = _find_first_file(stibo_dir, "Product") if stibo_dir.is_dir() else None
    if stibo_path is None:
        p = Path("STIBO/extract_stibo_all_products.xlsx")
        stibo_path = p if p.exists() else None
    return erp_path, ct_path, stibo_path


def get_input_files_hash(date_folder: str = "2302", market: str = "Ekofisk", erp_name: str = "Jeeves") -> str:
    """Calculate combined hash of all input files (ERP, CT, STIBO from dated folder or root)."""
    jeves_path, ct_path, stibo_path = _resolve_product_paths(date_folder, market, erp_name)
    input_files = [str(p) for p in (jeves_path, ct_path, stibo_path) if p is not None and p.exists()]
    if len(input_files) < 3:
        return None
    hashes = []
    for file_path in input_files:
        file_hash = get_file_hash(file_path)
        if file_hash:
            hashes.append(f"{file_path}:{file_hash}")
        else:
            return None
    combined = "|".join(hashes)
    return hashlib.md5(combined.encode()).hexdigest()

def find_existing_output_files(output_dir: Path) -> dict:
    """Find existing output files in output_dir."""
    files = {}
    range_files = list(output_dir.glob("Range_Reconciliation_*.xlsx"))
    if range_files:
        files["range"] = max(range_files, key=lambda x: x.stat().st_mtime)
    return files


def save_hash_info(input_hash: str, output_file: Path, output_dir: Path) -> None:
    """Save input hash with output file name in output_dir."""
    hash_file = output_dir / ".reconciliation_hash.json"
    hash_info = {"input_hash": input_hash, "output_file": str(output_file)}
    with open(hash_file, "w") as f:
        json.dump(hash_info, f, indent=2)


def load_hash_info(output_dir: Path) -> dict | None:
    """Load previous hash from output_dir."""
    hash_file = output_dir / ".reconciliation_hash.json"
    if hash_file.exists():
        try:
            with open(hash_file, "r") as f:
                return json.load(f)
        except Exception:
            return None
    return None


def main(
    date_folder: str = "2302",
    output_dir: Path | None = None,
    write_range_file: bool = True,
    market: str = "Ekofisk",
) -> pl.DataFrame:
    """Run Product (Range) reconciliation. Sources: ERP/{market}/{date}/, CT/{date}/, STIBO/{date}/.
    Returns the Product reconciliation DataFrame.
    If write_range_file=True, also writes Range_Reconciliation_*.xlsx to output_dir."""
    from market_config import get_erp_name
    erp_name = get_erp_name(market)

    date_folder = date_folder.strip()
    out = output_dir if output_dir is not None else Path(".")
    out.mkdir(parents=True, exist_ok=True)
    print(f"  Chargement des sources (market={market}, ERP={erp_name})...")

    erp_path, ct_path, stibo_path = _resolve_product_paths(date_folder, market, erp_name)
    if erp_path is None:
        raise FileNotFoundError(
            f"ERP Product file not found in ERP/{market}/{date_folder}/ nor ERP/{market}/."
        )
    if ct_path is None:
        raise FileNotFoundError(f"CT Product file not found in CT/{date_folder}/ nor CT/.")
    if stibo_path is None:
        raise FileNotFoundError(
            f"STIBO Product file not found in STIBO/{date_folder}/ nor STIBO/extract_stibo_all_products.xlsx."
        )

    print("  Lecture:")
    erp_df = _load_erp_product_data(str(erp_path), erp_name)
    print(f"    - {erp_name}: {len(erp_df)} produits")
    ct_df = load_ct_data(str(ct_path))
    print(f"    - CT:     {len(ct_df)} produits")
    stibo_df = load_stibo_data(str(stibo_path))
    print(f"    - STIBO:  {len(stibo_df)} produits")

    print(f"  Réconciliation (présence CT / {erp_name} / STIBO)...")
    range_reconciliation = create_range_reconciliation(erp_df, ct_df, stibo_df, erp_name)
    print(f"  Total produits uniques: {len(range_reconciliation)}")

    if write_range_file:
        current_input_hash = get_input_files_hash(date_folder, market, erp_name)
        previous_hash_info = load_hash_info(out)
        if current_input_hash and previous_hash_info and previous_hash_info.get("input_hash") == current_input_hash:
            existing_files = find_existing_output_files(out)
            output_file_range = existing_files.get("range", out / "Range_Reconciliation.xlsx")
            print("  Fichiers sources inchangés -> écrasement du fichier existant")
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file_range = out / f"Range_Reconciliation_{timestamp}.xlsx"
            if current_input_hash:
                print("  Fichiers sources modifiés -> nouveau fichier avec horodatage")
            else:
                print("  Attention: certains fichiers sources manquants -> nouveau fichier")
        range_reconciliation.write_excel(output_file_range)
        print(f"  -> Écrit: {output_file_range}")
        if current_input_hash:
            save_hash_info(current_input_hash, output_file_range, out)
    else:
        print(f"  -> Product data prêt pour intégration dans Reconciliation_{market}.xlsx")

    return range_reconciliation


if __name__ == "__main__":
    main(date_folder="2302", market="Ekofisk", write_range_file=True)
