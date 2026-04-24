"""
Invoice and Ordering-Shipping reconciliation (STIBO, CT, local ERP).
Compares Vendor and Customer across sources. Supports any market defined in markets.json.
Output: Reconciliation_{market}.xlsx with 5 sheets.
ERP files are read from ERP/{market}/{date}/.
"""
import polars as pl
from pathlib import Path
from openpyxl import load_workbook, Workbook

# Paths
STIBO_DIR = Path("STIBO")
CT_DIR = Path("CT")
ERP_DIR = Path("ERP")
JEEVES_DIR = Path("JEEVES")

# File patterns
CT_VENDOR_NEEDLE = "Vendor"
CT_CUSTOMER_NEEDLE = "Customer"
ERP_VENDOR_NEEDLE = "Vendor"
ERP_CUSTOMER_NEEDLE = "Customer"

# CT: sheet names and data start
CT_SHEET_INVOICE = "Invoice"
CT_SHEET_ORDERING = "OrderingShipping"
CT_COL_C = 3
CT_VENDOR_OS_COL = 4  # Column D for Vendor Ordering-Shipping
CT_CUSTOMER_OS_COL = 4  # Column D for Customer Ordering-Shipping (D8+)
CT_FIRST_ROW = 8

# Jeeves: Customer = headers row 2, data row 3+, column A. Vendor = headers row 1, column "SUVC -Invoice"
JEEVES_CUSTOMER_HEADER_ROW = 2
JEEVES_CUSTOMER_DATA_ROW = 3
JEEVES_VENDOR_HEADER_ROW = 1
JEEVES_VENDOR_DATA_ROW = 2
JEEVES_VENDOR_INVOICE_COL = "SUVC -Invoice"

KEY_COL = "Code"


def find_first_file(directory: Path, needle: str, market: str | None = None) -> Path | None:
    """First file in directory whose name contains needle (and market if given)."""
    if not directory.exists():
        return None
    needle_l = needle.lower()
    market_l = (market or "").lower()
    for f in sorted(directory.iterdir()):
        if not f.is_file():
            continue
        name_l = f.name.lower()
        if needle_l not in name_l:
            continue
        if market_l and market_l not in name_l:
            continue
        return f
    return None


def load_ct_column(path: Path, sheet_name: str, col: int = CT_COL_C, start_row: int = CT_FIRST_ROW) -> pl.DataFrame:
    """Load one column from Excel sheet from (col, start_row) until first empty."""
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not in {path.name}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    values = []
    for row in range(start_row, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            break
        values.append(str(v).strip())
    wb.close()
    return pl.DataFrame({KEY_COL: values})


def load_stibo_extract_column(extract_path: Path, sheet_name: str) -> pl.DataFrame:
    """Load the single data column from a STIBO extract file."""
    result = pl.read_excel(extract_path, sheet_name=sheet_name, raise_if_empty=False)
    if isinstance(result, dict):
        df = result.get(sheet_name) or list(result.values())[0]
    else:
        df = result
    if df is None or df.height == 0 or df.width == 0:
        return pl.DataFrame({KEY_COL: []})
    col = df.columns[0]
    return df.rename({col: KEY_COL}).select(pl.col(KEY_COL))


# STIBO: dated folder STIBO/{date}/, files e.g. Invoice_Vendors_{date}.xlsx
STIBO_OS_VENDORS_COL = "SUVC Ordering/Shipping"
STIBO_CUSTOMER_INVOICE_COL = "Invoice Customer Code"
STIBO_VENDOR_EXTRACT_ROOT = STIBO_DIR / "Vendor_extracts_STIBO.xlsx"
STIBO_CUSTOMER_EXTRACT_ROOT = STIBO_DIR / "Customer_extracts_STIBO.xlsx"


def _stibo_header_col(ws, header_aliases: tuple[str, ...]) -> int:
    """Return 1-based column index where normalized header matches one of header_aliases, else 1."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for i, h in enumerate(headers):
        if h is None:
            continue
        n = str(h).strip().lower().replace(" ", "")
        for alias in header_aliases:
            if n == alias:
                return i + 1
    return 1


def load_stibo_vendor_invoice_2302(path: Path) -> pl.DataFrame:
    """Load STIBO Vendor Invoice codes from Invoice_Vendors_2302.xlsx (1st col or 'SUVC Invoice')."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    col_idx = _stibo_header_col(ws, ("suvcinvoice", "suvc-invoice"))
    values = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(v)
    wb.close()
    return pl.DataFrame({KEY_COL: values})


def load_stibo_os_vendors(path: Path) -> pl.DataFrame:
    """Load STIBO Vendor OS codes from file (e.g. OS_Vendors_2302.xlsx), column 'SUVC Ordering/Shipping'."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def norm(s: str | None) -> str:
        if s is None:
            return ""
        return str(s).strip().lower().replace(" ", "")

    col_idx = None
    for i, h in enumerate(headers):
        if norm(h) == "suvcordering/shipping":
            col_idx = i + 1
            break
    if col_idx is None:
        wb.close()
        raise ValueError(
            f"Column '{STIBO_OS_VENDORS_COL}' not found in {path.name}. Available: {[h for h in headers if h]}"
        )
    values = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(v)
    wb.close()
    return pl.DataFrame({KEY_COL: values})


def load_stibo_os_customers(path: Path) -> pl.DataFrame:
    """Load STIBO Customer OS codes from file (e.g. OS_Customers_2304.xlsx).

    Expected a column like 'Customer Code Ordering / Shipping' (or similar) on row 1.
    Keeps leading zeros by reading values as-is (strings) and later normalizing via _normalize_os_codes.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def norm(s: str | None) -> str:
        if s is None:
            return ""
        return str(s).strip().lower().replace(" ", "")

    col_idx = None
    # More tolerant: accept headers like "Ordering Cust" as well
    for i, h in enumerate(headers):
        nh = norm(h)
        if (
            ("customercode" in nh and "ordering" in nh)
            or ("ordering/shipping" in nh and "customer" in nh)
            or (("ordering" in nh) and ("cust" in nh))
            or nh in ("orderingcust", "orderingcust.", "orderingcustcode")
        ):
            col_idx = i + 1
            break
    if col_idx is None:
        # Fallback: assume first column contains the codes (as in your extract screenshot)
        col_idx = 1

    values = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(v)
    wb.close()
    return pl.DataFrame({KEY_COL: values})


def load_stibo_customer_invoice(path: Path) -> pl.DataFrame:
    """Load STIBO Customer Invoice codes from file (e.g. Invoice_Customer_2302.xlsx), column 'Invoice Customer Code' (Q)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def norm(s: str | None) -> str:
        if s is None:
            return ""
        return str(s).strip().lower().replace(" ", "")

    col_idx = None
    for i, h in enumerate(headers):
        if norm(h) == "invoicecustomercode":
            col_idx = i + 1
            break
    if col_idx is None:
        wb.close()
        raise ValueError(
            f"Column '{STIBO_CUSTOMER_INVOICE_COL}' not found in {path.name}. Available: {[h for h in headers if h]}"
        )
    values = []
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(str(v).strip())
    wb.close()
    return pl.DataFrame({KEY_COL: pl.Series(values).cast(pl.Utf8)})


def load_jeves_vendor_invoice(path: Path) -> pl.DataFrame:
    """JEEVES Vendor: headers row 1, data row 2+, column 'SUVC -Invoice'."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=JEEVES_VENDOR_HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    # Try possible column names (spacing varies: "SUVC - Invoice ", "SUVC -Invoice", etc.)
    col_idx = None
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_clean = str(h).strip().lower().replace(" ", "")
        if h_clean == "suvc-invoice":
            col_idx = i + 1
            break
    if col_idx is None:
        wb.close()
        raise ValueError(f"Column '{JEEVES_VENDOR_INVOICE_COL}' not in JEEVES Vendor. Available: {headers}")
    values = []
    for row in range(JEEVES_VENDOR_DATA_ROW, ws.max_row + 1):
        v = ws.cell(row=row, column=col_idx).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(v)
    wb.close()
    return pl.DataFrame({KEY_COL: values})


# JEEVES Vendor OS: same file as Vendor Invoice, sheet "ORDERSHIPPING", column A
JEEVES_VENDOR_OS_SHEETS = ("ORDERSHIPPING", "ODERSHIPPING", "OrderingShipping", "ORDERINGSHIPPING")
JEEVES_VENDOR_OS_HEADER_ROW = 1
JEEVES_VENDOR_OS_DATA_ROW = 2


def load_jeves_vendor_ordering(path: Path) -> pl.DataFrame:
    """JEEVES Vendor OS: sheet 'ORDERSHIPPING', column A, data from row 2."""
    wb = load_workbook(path, data_only=True)
    ws = None
    for name in JEEVES_VENDOR_OS_SHEETS:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        raise ValueError(
            f"Sheet for Vendor OS not found in {path.name}. Tried: {JEEVES_VENDOR_OS_SHEETS}. Available: {wb.sheetnames}"
        )
    values = []
    for row in range(JEEVES_VENDOR_OS_DATA_ROW, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(str(v).strip())
    wb.close()
    return pl.DataFrame({KEY_COL: pl.Series(values).cast(pl.Utf8)})


# JEEVES Customer Invoice: sheet "INVOICECUSTOMER", column A, headers row 2, data from row 3
JEEVES_CUSTOMER_INVOICE_SHEET = "INVOICECUSTOMER"
# JEEVES Customer OS: sheet "ORDERSHIPPING", column A from row 3
JEEVES_CUSTOMER_OS_SHEETS = ("ORDERSHIPPING", "OrderShipping", "ORDERINGSHIPPING")
JEEVES_CUSTOMER_OS_DATA_ROW = 3


def load_jeves_customer_invoice(path: Path) -> pl.DataFrame:
    """JEEVES Customer Invoice: sheet 'INVOICECUSTOMER', column A, headers row 2, data from row 3."""
    wb = load_workbook(path, data_only=True)
    if JEEVES_CUSTOMER_INVOICE_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{JEEVES_CUSTOMER_INVOICE_SHEET}' not in {path.name}. Available: {wb.sheetnames}"
        )
    ws = wb[JEEVES_CUSTOMER_INVOICE_SHEET]
    values = []
    for row in range(JEEVES_CUSTOMER_DATA_ROW, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        values.append(str(v).strip())
    wb.close()
    return pl.DataFrame({KEY_COL: pl.Series(values).cast(pl.Utf8)})


def _jeves_os_customer_code_raw(val) -> str | None:
    """Preserve JEEVES Customer OS code as string; keep leading zeros. If Excel returns int 5, pad to '0005'."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = int(float(val))
        return f"{n:04d}" if 0 <= n < 10000 else str(n)
    s = str(val).strip()
    return s if s else None


def load_jeves_customer_ordering(path: Path) -> pl.DataFrame:
    """JEEVES Customer OS: sheet 'ORDERSHIPPING', column A from row 3. Preserves leading zeros (e.g. 0005)."""
    wb = load_workbook(path, data_only=True)
    ws = None
    for name in JEEVES_CUSTOMER_OS_SHEETS:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        raise ValueError(
            f"Sheet for Customer OS not found in {path.name}. Tried: {JEEVES_CUSTOMER_OS_SHEETS}. Available: {wb.sheetnames}"
        )
    values = []
    for row in range(JEEVES_CUSTOMER_OS_DATA_ROW, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        code = _jeves_os_customer_code_raw(v)
        if code is not None:
            values.append(code)
    wb.close()
    return pl.DataFrame({KEY_COL: pl.Series(values).cast(pl.Utf8)})


def _normalize(df: pl.DataFrame) -> pl.DataFrame:
    return df.with_columns(pl.col(KEY_COL).cast(pl.Utf8).str.strip_chars()).filter(
        pl.col(KEY_COL).is_not_null() & (pl.col(KEY_COL) != "")
    )


def _os_customer_code_to_str(val) -> str:
    """Convert value to string for OS Customer code; preserve leading zeros (e.g. 5 -> '0005')."""
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = int(float(val))
        if 0 <= n < 10000:
            return f"{n:04d}"
        return str(n)
    s = str(val).strip()
    # String that looks like an integer: zero-pad to 4 digits for consistent comparison
    if s.isdigit():
        n = int(s)
        if 0 <= n < 10000:
            return f"{n:04d}"
    return s


def _normalize_os_codes(df: pl.DataFrame) -> pl.DataFrame:
    """Compare OS codes as strings; preserve leading zeros (e.g. 5 -> '0005'). Used for Vendor OS and Customer OS."""
    if df.height == 0:
        return df
    codes = [_os_customer_code_to_str(v) for v in df[KEY_COL].to_list()]
    return pl.DataFrame({KEY_COL: pl.Series(codes).cast(pl.Utf8)}).filter(
        pl.col(KEY_COL).is_not_null() & (pl.col(KEY_COL) != "")
    )


def build_reconciliation(
    stibo_vendor: pl.DataFrame,
    stibo_customer: pl.DataFrame,
    ct_vendor: pl.DataFrame,
    ct_customer: pl.DataFrame,
    erp_vendor: pl.DataFrame,
    erp_customer: pl.DataFrame,
    erp_name: str,
) -> pl.DataFrame:
    """Build reconciliation table: all unique codes with X per source."""
    key = KEY_COL
    sets = {
        "stibo_v": set(stibo_vendor[key].to_list()),
        "stibo_c": set(stibo_customer[key].to_list()),
        "ct_v": set(ct_vendor[key].to_list()),
        "ct_c": set(ct_customer[key].to_list()),
        "erp_v": set(erp_vendor[key].to_list()),
        "erp_c": set(erp_customer[key].to_list()),
    }
    all_codes = sorted(set.union(*sets.values()))
    return pl.DataFrame({
        key: all_codes,
        "STIBO_Vendor": ["X" if c in sets["stibo_v"] else "" for c in all_codes],
        "STIBO_Customer": ["X" if c in sets["stibo_c"] else "" for c in all_codes],
        "CT_Vendor": ["X" if c in sets["ct_v"] else "" for c in all_codes],
        "CT_Customer": ["X" if c in sets["ct_c"] else "" for c in all_codes],
        f"{erp_name}_Vendor": ["X" if c in sets["erp_v"] else "" for c in all_codes],
        f"{erp_name}_Customer": ["X" if c in sets["erp_c"] else "" for c in all_codes],
    })


# Sheet names for the single output file (5 tabs)
SHEET_PRODUCT = "Product"
SHEET_VENDOR_INVOICE = "Vendor Invoice"
SHEET_VENDOR_OS = "Vendor OS"
SHEET_CUSTOMER_INVOICE = "Customer Invoice"
SHEET_CUSTOMER_OS = "Customer OS"


def _sheet_from_full(full_df: pl.DataFrame, vendor: bool, erp_name: str) -> pl.DataFrame:
    """Extract Vendor or Customer view: rows where at least one source has X, Code + 3 source columns."""
    src_cols = (
        ["STIBO_Vendor", "CT_Vendor", f"{erp_name}_Vendor"]
        if vendor
        else ["STIBO_Customer", "CT_Customer", f"{erp_name}_Customer"]
    )
    available_src = [c for c in src_cols if c in full_df.columns]
    if not available_src:
        out_cols = [KEY_COL] if KEY_COL in full_df.columns else full_df.columns[:1]
        return full_df.select(out_cols)
    mask = pl.lit(False)
    for c in available_src:
        mask = mask | (pl.col(c) == "X")
    filtered = full_df.filter(mask)
    out_cols = [KEY_COL] + available_src
    return filtered.select([c for c in out_cols if c in filtered.columns])


def write_reconciliation_excel_5_tabs(
    path: Path,
    rec_invoice: pl.DataFrame,
    rec_ordering: pl.DataFrame,
    product_df: pl.DataFrame | None = None,
    erp_name: str = "ERP",
) -> None:
    """Write Reconciliation_{market}.xlsx with 5 sheets: Product, Vendor Invoice, Vendor OS, Customer Invoice, Customer OS."""
    wb = Workbook()
    del wb["Sheet"]

    empty_product = pl.DataFrame(
        {"ProductCode": [], "CT": [], erp_name: [], "STIBO": [], "Absent_from": []}
    )
    sheets = [
        (SHEET_PRODUCT, product_df if product_df is not None and product_df.height > 0 else empty_product),
        (SHEET_VENDOR_INVOICE, _sheet_from_full(rec_invoice, vendor=True, erp_name=erp_name)),
        (SHEET_VENDOR_OS, _sheet_from_full(rec_ordering, vendor=True, erp_name=erp_name)),
        (SHEET_CUSTOMER_INVOICE, _sheet_from_full(rec_invoice, vendor=False, erp_name=erp_name)),
        (SHEET_CUSTOMER_OS, _sheet_from_full(rec_ordering, vendor=False, erp_name=erp_name)),
    ]
    for sheet_name, df in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(df.columns)
        for row in df.iter_rows(named=False):
            ws.append(list(row))
        # Customer OS: force Code column (A) as text so Excel keeps leading zeros (e.g. 0005)
        if sheet_name == SHEET_CUSTOMER_OS and df.height > 0:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=1).number_format = "@"
    wb.save(path)


def _load_erp_vendor_invoice(path: Path, erp_name: str) -> pl.DataFrame:
    if erp_name.lower() == "jeeves":
        return load_jeves_vendor_invoice(path)
    raise NotImplementedError(f"No Vendor Invoice loader for ERP '{erp_name}'.")


def _load_erp_vendor_ordering(path: Path, erp_name: str) -> pl.DataFrame:
    if erp_name.lower() == "jeeves":
        return load_jeves_vendor_ordering(path)
    raise NotImplementedError(f"No Vendor OS loader for ERP '{erp_name}'.")


def _load_erp_customer_invoice(path: Path, erp_name: str) -> pl.DataFrame:
    if erp_name.lower() == "jeeves":
        return load_jeves_customer_invoice(path)
    raise NotImplementedError(f"No Customer Invoice loader for ERP '{erp_name}'.")


def _load_erp_customer_ordering(path: Path, erp_name: str) -> pl.DataFrame:
    if erp_name.lower() == "jeeves":
        return load_jeves_customer_ordering(path)
    raise NotImplementedError(f"No Customer OS loader for ERP '{erp_name}'.")


def run_invoice_ordering_reconciliation(
    market: str,
    output_dir: Path,
    product_df: pl.DataFrame | None = None,
    date_folder: str = "2302",
) -> Path:
    """Run Invoice + Ordering-Shipping reconciliation for one market. Returns output path.
    Sources: STIBO/{date_folder}/, CT/{date_folder}/, ERP/{market}/{date_folder}/."""
    from market_config import get_erp_name
    erp_name = get_erp_name(market)

    market_filter = market if market else None
    date_folder = date_folder.strip()

    # Dated dirs: prefer dated subdir, fallback to root
    stibo_date_dir = STIBO_DIR / date_folder
    ct_search_dir = CT_DIR / date_folder if (CT_DIR / date_folder).is_dir() else CT_DIR
    erp_base = ERP_DIR / erp_name
    # Backward-compatible: for Jeeves, support legacy layout JEEVES/{date}/ when ERP/ is not used.
    if erp_name.lower() == "jeeves" and not erp_base.exists():
        erp_base = JEEVES_DIR
    erp_search_dir = erp_base / date_folder if (erp_base / date_folder).is_dir() else erp_base

    # STIBO: files in STIBO/{date}/ e.g. Invoice_Vendors_2302.xlsx
    vendor_extract = STIBO_VENDOR_EXTRACT_ROOT
    customer_extract = STIBO_CUSTOMER_EXTRACT_ROOT
    stibo_inv_vendors = stibo_date_dir / f"Invoice_Vendors_{date_folder}.xlsx"
    stibo_os_vendors = stibo_date_dir / f"OS_Vendors_{date_folder}.xlsx"
    stibo_inv_customers = (
        stibo_date_dir / f"Invoice_Customers_{date_folder}.xlsx",
        stibo_date_dir / f"Invoice_Customer_{date_folder}.xlsx",
    )
    stibo_os_customers = stibo_date_dir / f"OS_Customers_{date_folder}.xlsx"

    if stibo_inv_vendors.exists():
        stibo_vendor_inv = _normalize(load_stibo_vendor_invoice_2302(stibo_inv_vendors))
    elif vendor_extract.exists():
        stibo_vendor_inv = _normalize(load_stibo_extract_column(vendor_extract, "Invoice"))
    else:
        raise FileNotFoundError(
            f"STIBO Vendor Invoice: not found {stibo_inv_vendors} nor {vendor_extract}."
        )
    if stibo_os_vendors.exists():
        stibo_vendor_ord = _normalize(load_stibo_os_vendors(stibo_os_vendors))
    elif vendor_extract.exists():
        stibo_vendor_ord = _normalize(load_stibo_extract_column(vendor_extract, "Ordering-Shipping"))
    else:
        stibo_vendor_ord = _normalize(pl.DataFrame({KEY_COL: []}))

    stibo_customer_inv_file = None
    for p in stibo_inv_customers:
        if p.exists():
            stibo_customer_inv_file = p
            break
    if stibo_customer_inv_file is not None:
        stibo_customer_inv = _normalize(load_stibo_customer_invoice(stibo_customer_inv_file))
    elif customer_extract.exists():
        stibo_customer_inv = _normalize(load_stibo_extract_column(customer_extract, "Invoice"))
    else:
        raise FileNotFoundError(
            f"STIBO Customer Invoice: not found {stibo_inv_customers[0]} nor {customer_extract}."
        )
    if stibo_os_customers.exists():
        stibo_customer_ord = _normalize(load_stibo_os_customers(stibo_os_customers))
    elif customer_extract.exists():
        stibo_customer_ord = _normalize(load_stibo_extract_column(customer_extract, "Ordering-Shipping"))
    else:
        stibo_customer_ord = _normalize(pl.DataFrame({KEY_COL: []}))

    # CT: search in dated folder or root
    ct_vendor_file = find_first_file(ct_search_dir, CT_VENDOR_NEEDLE, market_filter)
    ct_customer_file = find_first_file(ct_search_dir, CT_CUSTOMER_NEEDLE, market_filter)
    if not ct_vendor_file:
        raise FileNotFoundError(f"CT Vendor file not found in {ct_search_dir.absolute()} (market={market or 'any'}).")
    if not ct_customer_file:
        raise FileNotFoundError(f"CT Customer file not found in {ct_search_dir.absolute()} (market={market or 'any'}).")

    ct_vendor_inv = _normalize(load_ct_column(ct_vendor_file, CT_SHEET_INVOICE))
    ct_vendor_ord = _normalize(load_ct_column(ct_vendor_file, CT_SHEET_ORDERING, col=CT_VENDOR_OS_COL))
    ct_customer_inv = _normalize(load_ct_column(ct_customer_file, CT_SHEET_INVOICE))
    ct_customer_ord = _normalize(load_ct_column(ct_customer_file, CT_SHEET_ORDERING, col=CT_CUSTOMER_OS_COL))

    # ERP: search in ERP/{market}/{date}/ or ERP/{market}/
    erp_vendor_file = find_first_file(erp_search_dir, ERP_VENDOR_NEEDLE, market=None)
    erp_customer_file = find_first_file(erp_search_dir, ERP_CUSTOMER_NEEDLE, market=None)
    if not erp_vendor_file:
        raise FileNotFoundError(
            f"{erp_name} Vendor file not found in {erp_search_dir.absolute()}. "
            f"Expected a file with 'Vendor' in name."
        )
    if not erp_customer_file:
        raise FileNotFoundError(
            f"{erp_name} Customer file not found in {erp_search_dir.absolute()}. "
            f"Expected a file with 'Customer' in name."
        )

    erp_vendor_inv = _normalize(_load_erp_vendor_invoice(erp_vendor_file, erp_name))
    erp_customer_inv = _normalize(_load_erp_customer_invoice(erp_customer_file, erp_name))
    erp_vendor_ord = _normalize(_load_erp_vendor_ordering(erp_vendor_file, erp_name))
    erp_customer_ord = _normalize(_load_erp_customer_ordering(erp_customer_file, erp_name))

    # OS (Vendor + Customer): compare codes as strings with leading zeros (e.g. "0005" not "5")
    stibo_vendor_ord = _normalize_os_codes(stibo_vendor_ord)
    ct_vendor_ord = _normalize_os_codes(ct_vendor_ord)
    erp_vendor_ord = _normalize_os_codes(erp_vendor_ord)
    stibo_customer_ord = _normalize_os_codes(stibo_customer_ord)
    ct_customer_ord = _normalize_os_codes(ct_customer_ord)
    erp_customer_ord = _normalize_os_codes(erp_customer_ord)

    rec_invoice = build_reconciliation(
        stibo_vendor_inv, stibo_customer_inv, ct_vendor_inv, ct_customer_inv,
        erp_vendor_inv, erp_customer_inv, erp_name,
    )
    rec_ordering = build_reconciliation(
        stibo_vendor_ord, stibo_customer_ord, ct_vendor_ord, ct_customer_ord,
        erp_vendor_ord, erp_customer_ord, erp_name,
    )

    out_path = output_dir / f"Reconciliation_{market}.xlsx"
    write_reconciliation_excel_5_tabs(out_path, rec_invoice, rec_ordering, product_df=product_df, erp_name=erp_name)
    print(f"  Market: {market} (ERP: {erp_name})")
    print(f"  Fichiers lus: STIBO (Vendor/Customer), CT ({ct_vendor_file.name} / {ct_customer_file.name}), {erp_name} ({erp_vendor_file.name} / {erp_customer_file.name})")
    print("  Onglets generes: Product | Vendor Invoice | Vendor OS | Customer Invoice | Customer OS")
    print(f"  Lignes: Invoice={rec_invoice.height}, Ordering-Shipping={rec_ordering.height}")
    print(f"  -> ecrit: {out_path}")
    return out_path


def main(date_folder: str = "2302") -> None:
    run_invoice_ordering_reconciliation("Ekofisk", Path("."), date_folder=date_folder)


if __name__ == "__main__":
    main()
