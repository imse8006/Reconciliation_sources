"""
Extract Invoice columns from STIBO sources for Vendor and Customer.
Produces centralization files: Vendor_extracts_STIBO.xlsx and Customer_extracts_STIBO.xlsx.
Each file has sheets: "Invoice" (extracted column), "Ordering/Shipping" (placeholder for later).
"""
import polars as pl
from pathlib import Path
from openpyxl import Workbook

# Column names to extract
VENDOR_INVOICE_COL = "SUVC Invoice"
CUSTOMER_INVOICE_COL = "Invoice Customer Code"

# File patterns (STIBO directory)
SOURCE_DIR = Path("STIBO")
EXCEL_VENDOR_PATTERN = "excel*2026*02*18*.xlsx"
STIBO_CUSTOMER_PATTERN = "stibo-eu-invoice-customers*.xlsx"


def find_first_file(directory: Path, pattern: str, recursive: bool = True) -> Path | None:
    """Find first file matching glob pattern under directory (optionally recursive)."""
    pattern_rec = f"**/{pattern}" if recursive else pattern
    matches = sorted(directory.glob(pattern_rec), key=lambda p: (len(p.parts), str(p)))
    return matches[0] if matches else None


def load_column_from_excel(path: Path, column_name: str, first_sheet: bool = True) -> pl.DataFrame:
    """Load a single column from an Excel file. Uses first sheet if first_sheet=True."""
    # Read Excel file (may return dict if multiple sheets, or DataFrame if single sheet)
    result = pl.read_excel(path)
    
    # Handle case where read_excel returns a dict (multiple sheets)
    if isinstance(result, dict):
        df = list(result.values())[0]  # Take first sheet
    else:
        df = result
    
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found. Available: {df.columns}")
    return df.select(pl.col(column_name))


def write_excel_two_sheets(
    out_path: Path,
    invoice_df: pl.DataFrame,
    ordering_sheet_name: str = "Ordering-Shipping",
    invoice_sheet_name: str = "Invoice",
) -> None:
    """Write an Excel file with two sheets: Invoice (with data) and Ordering-Shipping (empty placeholder)."""
    wb = Workbook()
    # Remove default sheet and create our two sheets
    del wb["Sheet"]

    ws_inv = wb.create_sheet(invoice_sheet_name, 0)
    headers = invoice_df.columns
    ws_inv.append(list(headers))
    for row in invoice_df.iter_rows(named=False):
        ws_inv.append(list(row))

    ws_ord = wb.create_sheet(ordering_sheet_name, 1)
    ws_ord.append([])  # placeholder

    wb.save(out_path)


def main() -> None:
    source_dir = SOURCE_DIR

    # --- Vendor: excel-2026-02-18....xlsx -> SUVC Invoice
    vendor_file = find_first_file(source_dir, EXCEL_VENDOR_PATTERN)
    if not vendor_file:
        raise FileNotFoundError(
            f"No file matching '{EXCEL_VENDOR_PATTERN}' in {source_dir.absolute()}. "
            "Place the Vendor Excel file in the project root."
        )
    vendor_invoice = load_column_from_excel(vendor_file, VENDOR_INVOICE_COL)
    vendor_out = source_dir / "Vendor_extracts_STIBO.xlsx"
    write_excel_two_sheets(vendor_out, vendor_invoice)
    print(f"Vendor: {vendor_invoice.height} rows -> {vendor_out}")

    # --- Customer: stibo-eu-invoice-customers....xlsx -> Invoice Customer Code
    customer_file = find_first_file(source_dir, STIBO_CUSTOMER_PATTERN)
    if not customer_file:
        raise FileNotFoundError(
            f"No file matching '{STIBO_CUSTOMER_PATTERN}' in {source_dir.absolute()}. "
            "Place the STIBO customer invoice file in the project root."
        )
    customer_invoice = load_column_from_excel(customer_file, CUSTOMER_INVOICE_COL)
    customer_out = source_dir / "Customer_extracts_STIBO.xlsx"
    write_excel_two_sheets(customer_out, customer_invoice)
    print(f"Customer: {customer_invoice.height} rows -> {customer_out}")


if __name__ == "__main__":
    main()
